"""
Aror University Transcript PDF -> Excel (exact layout replica)
Usage: python3 transcript_to_excel.py input.pdf output.xlsx
"""
import sys
import re
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

LEFT_BOUNDS = [('code', 0, 63), ('title', 63, 190), ('ch', 190, 214),
               ('gp', 214, 237), ('mks', 237, 262), ('grd', 262, 306)]
RIGHT_BOUNDS = [('code', 306, 358), ('title', 358, 486), ('ch', 486, 509),
                ('gp', 509, 532), ('mks', 532, 557), ('grd', 557, 612)]

THIN = Side(style='thin', color='000000')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BOLD = Font(bold=True, name='Arial', size=9)
NORMAL = Font(name='Arial', size=9)
CENTER = Alignment(horizontal='center', vertical='center')
LEFT_AL = Alignment(horizontal='left', vertical='center')


def cluster_rows(words, tol=3):
    words = sorted(words, key=lambda w: w['top'])
    rows, cur, cur_top = [], [], None
    for w in words:
        if cur_top is None or abs(w['top'] - cur_top) <= tol:
            cur.append(w)
            cur_top = w['top'] if cur_top is None else cur_top
        else:
            rows.append(cur)
            cur, cur_top = [w], w['top']
    if cur:
        rows.append(cur)
    return rows


def classify(x0, side):
    bounds = LEFT_BOUNDS if side == 'L' else RIGHT_BOUNDS
    for name, lo, hi in bounds:
        if lo <= x0 < hi:
            return name
    return None


def bin_row(row_words):
    """Split a clustered row into left-half and right-half word lists."""
    left = [w for w in row_words if w['x0'] < 306]
    right = [w for w in row_words if w['x0'] >= 306]
    return left, right


def words_to_fields(words, side):
    fields = {'code': [], 'title': [], 'ch': [], 'gp': [], 'mks': [], 'grd': []}
    for w in sorted(words, key=lambda w: w['x0']):
        col = classify(w['x0'], side)
        if col:
            fields[col].append(w['text'])
    return {k: ' '.join(v) for k, v in fields.items()}


def is_course_row(fields):
    return bool(fields['code']) and bool(re.match(r'^\d', fields['ch'] or '')) is False and (
        fields['ch'] or fields['gp'] or fields['mks'] or fields['grd'])


def is_semester_header(words, side_words_all_in_code_title=True):
    text = ' '.join(w['text'] for w in words)
    return bool(re.search(r'(Semester|Summer\s+\d{4})', text)) and not re.search(r'\d\.\d\d', text)


def parse_header_block(words):
    """Parse the top info block (top < ~176) into label: value pairs, left & right."""
    left_words = [w for w in words if w['x0'] < 306 and w['top'] < 176]
    right_words = [w for w in words if w['x0'] >= 306 and w['top'] < 176]
    rows = cluster_rows(left_words + right_words, tol=2)
    left_info, right_info = [], []
    for row in rows:
        left, right = bin_row(row)
        for group, bucket in ((left, left_info), (right, right_info)):
            if not group:
                continue
            group = sorted(group, key=lambda w: w['x0'])
            text = ' '.join(w['text'] for w in group)
            if ':' in text:
                label, _, val = text.partition(':')
                bucket.append((label.strip() + ':', val.strip()))
            else:
                bucket.append(('', text.strip()))
    return left_info, right_info


def parse_page1(page):
    words = page.extract_words()
    header_left, header_right = parse_header_block(words)

    table_words = [w for w in words if w['top'] >= 176]
    rows = cluster_rows(table_words, tol=3)

    left_blocks = []   # list of dicts: {'semester':..., 'courses':[...]}
    right_blocks = []
    cur_left_sem, cur_right_sem = None, None
    footer_words = []

    for row in rows:
        left, right = bin_row(row)
        row_text = ' '.join(w['text'] for w in row)

        if row_text.strip() in ('Course Code Title CH GP MKS GRD',) or \
           (('Course' in row_text and 'Code' in row_text and 'Title' in row_text)):
            continue  # header repeat row

        # Footer detection: CGPA / qualification / enrollment note / end marker
        if re.search(r'^CGPA', row_text.strip()) or 'has thus qualified' in row_text or \
           'is currently enrolled' in row_text or 'End of Transcript' in row_text or \
           'best of our knowledge' in row_text or 'degree certificate' in row_text or \
           'Controller of Examinations' in row_text or 'Vice Chancellor' in row_text:
            footer_words.extend(row)
            continue

        # Semester header detection (text only, no CH/GP numbers in fields)
        if left and is_semester_header(left):
            cur_left_sem = ' '.join(w['text'] for w in sorted(left, key=lambda w: w['x0']))
            left_blocks.append({'semester': cur_left_sem, 'courses': []})
            left = []
        if right and is_semester_header(right):
            cur_right_sem = ' '.join(w['text'] for w in sorted(right, key=lambda w: w['x0']))
            right_blocks.append({'semester': cur_right_sem, 'courses': []})
            right = []

        if left:
            f = words_to_fields(left, 'L')
            if f['code']:
                if not left_blocks:
                    left_blocks.append({'semester': '', 'courses': []})
                left_blocks[-1]['courses'].append(f)
        if right:
            f = words_to_fields(right, 'R')
            if f['code']:
                if not right_blocks:
                    right_blocks.append({'semester': '', 'courses': []})
                right_blocks[-1]['courses'].append(f)

    footer_words = sorted(footer_words, key=lambda w: (w['top'], w['x0']))
    footer_rows = cluster_rows(footer_words, tol=3)
    footer_lines = []
    cgpa_val = ''
    for row in footer_rows:
        row = sorted(row, key=lambda w: w['x0'])
        text = ' '.join(w['text'] for w in row)
        m = re.match(r'^CGPA\s+([\d.]+)', text.strip())
        if m:
            cgpa_val = m.group(1)
            continue
        footer_lines.append(text)

    return header_left, header_right, left_blocks, right_blocks, cgpa_val, footer_lines


def parse_grading_page(page):
    """Parse the Grading System table on the last page, if present."""
    words = page.extract_words()
    text = ' '.join(w['text'] for w in words)
    if 'Grading System' not in text:
        return None
    tbl = page.extract_table()
    return tbl


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_workbook(data, grading_table, out_path):
    header_left, header_right, left_blocks, right_blocks, cgpa_val, footer_lines = data
    wb = Workbook()
    ws = wb.active
    ws.title = 'Transcript'

    widths = [12, 24, 6, 6, 6, 6, 3, 12, 24, 6, 6, 6, 6]
    autosize(ws, widths)

    r = 1
    # ---- header info block ----
    for (llabel, lval), (rlabel, rval) in zip(header_left, header_right):
        ws.cell(r, 1, llabel).font = BOLD
        c = ws.cell(r, 3, lval); c.font = NORMAL
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        ws.cell(r, 8, rlabel).font = BOLD
        c = ws.cell(r, 10, rval); c.font = NORMAL
        ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=13)
        r += 1
    r += 1

    # ---- course table header ----
    def write_table_header(row):
        left_hdr = ['Course Code', 'Title', 'CH', 'GP', 'MKS', 'GRD']
        for i, h in enumerate(left_hdr):
            col = 1 if i == 0 else (2 if i == 1 else i + 1)
        ws.cell(row, 1, 'Course Code').font = BOLD
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1)
        ws.cell(row, 2, 'Title').font = BOLD
        ws.cell(row, 3, 'CH').font = BOLD
        ws.cell(row, 4, 'GP').font = BOLD
        ws.cell(row, 5, 'MKS').font = BOLD
        ws.cell(row, 6, 'GRD').font = BOLD
        ws.cell(row, 8, 'Course Code').font = BOLD
        ws.cell(row, 9, 'Title').font = BOLD
        ws.cell(row, 10, 'CH').font = BOLD
        ws.cell(row, 11, 'GP').font = BOLD
        ws.cell(row, 12, 'MKS').font = BOLD
        ws.cell(row, 13, 'GRD').font = BOLD
        for col in list(range(1, 7)) + list(range(8, 14)):
            cell = ws.cell(row, col)
            cell.alignment = CENTER
            cell.border = BOX

    write_table_header(r)
    header_row = r
    r += 1

    max_blocks = max(len(left_blocks), len(right_blocks))
    for i in range(max_blocks):
        lb = left_blocks[i] if i < len(left_blocks) else None
        rb = right_blocks[i] if i < len(right_blocks) else None

        # semester title rows
        if lb and lb['semester']:
            ws.cell(r, 1, lb['semester']).font = BOLD
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            ws.cell(r, 1).alignment = CENTER
        if rb and rb['semester']:
            ws.cell(r, 8, rb['semester']).font = BOLD
            ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=13)
            ws.cell(r, 8).alignment = CENTER
        sem_row = r
        r += 1

        lcount = len(lb['courses']) if lb else 0
        rcount = len(rb['courses']) if rb else 0
        rows_needed = max(lcount, rcount)
        for j in range(rows_needed):
            if lb and j < lcount:
                c = lb['courses'][j]
                ws.cell(r, 1, c['code']).font = NORMAL
                ws.cell(r, 2, c['title']).font = NORMAL
                ws.cell(r, 3, _num(c['ch'])).font = NORMAL
                ws.cell(r, 4, _num(c['gp'])).font = NORMAL
                ws.cell(r, 5, _num(c['mks'])).font = NORMAL
                ws.cell(r, 6, c['grd']).font = NORMAL
            if rb and j < rcount:
                c = rb['courses'][j]
                ws.cell(r, 8, c['code']).font = NORMAL
                ws.cell(r, 9, c['title']).font = NORMAL
                ws.cell(r, 10, _num(c['ch'])).font = NORMAL
                ws.cell(r, 11, _num(c['gp'])).font = NORMAL
                ws.cell(r, 12, _num(c['mks'])).font = NORMAL
                ws.cell(r, 13, c['grd']).font = NORMAL
            r += 1

    table_end_row = r - 1
    r += 1
    if cgpa_val:
        ws.cell(r, 12, 'CGPA').font = BOLD
        ws.cell(r, 13, float(cgpa_val)).font = BOLD
        r += 2

    for line in footer_lines:
        ws.cell(r, 1, line).font = NORMAL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
        ws.cell(r, 1).alignment = LEFT_AL
        r += 1

    r += 2
    ws.cell(r, 1, 'Controller of Examinations').font = BOLD
    ws.cell(r, 8, 'Vice Chancellor').font = BOLD

    # borders for the course-table area only
    for row in range(header_row, table_end_row + 1):
        for col in list(range(1, 7)) + list(range(8, 14)):
            ws.cell(row, col).border = BOX

    # ---- Grading system sheet ----
    if grading_table:
        ws2 = wb.create_sheet('Grading System')
        ws2.column_dimensions['A'].width = 10
        ws2.column_dimensions['B'].width = 20
        ws2.column_dimensions['C'].width = 14
        for ri, row in enumerate(grading_table, start=1):
            for ci, val in enumerate(row, start=1):
                cell = ws2.cell(ri, ci, val)
                cell.border = BOX
                cell.alignment = CENTER
                cell.font = BOLD if ri == 1 else NORMAL

    wb.save(out_path)


def _num(s):
    s = (s or '').strip()
    try:
        if s == '':
            return None
        return float(s) if '.' in s else int(s)
    except ValueError:
        return s


def main():
    if len(sys.argv) < 3:
        print('Usage: python3 transcript_to_excel.py input.pdf output.xlsx')
        sys.exit(1)
    in_pdf, out_xlsx = sys.argv[1], sys.argv[2]
    with pdfplumber.open(in_pdf) as pdf:
        page1 = pdf.pages[0]
        data = parse_page1(page1)
        grading_table = None
        for p in pdf.pages:
            gt = parse_grading_page(p)
            if gt:
                grading_table = gt
                break
    build_workbook(data, grading_table, out_xlsx)
    print(f'Saved: {out_xlsx}')


if __name__ == '__main__':
    main()
